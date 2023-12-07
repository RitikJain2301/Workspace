import openpyxl

file=openpyxl.load_workbook("C:/Users/Administrator/Desktop/test excels/multiple rows.xlsx")
# sheet=file.active
# print(sheet)

sh=file["Sheet1"]
maxRow=sh.max_row
maxCol=sh.max_column
for i in range(1,maxRow+1):
    for j in range(1,maxCol+1):
        print(sh.cell(i,j).value)

sh.cell(7,7).value="ritikkkkkk"
file.save("C:/Users/Administrator/Desktop/test excels/multiple rows.xlsx")