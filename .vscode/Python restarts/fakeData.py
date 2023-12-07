from openpyxl import Workbook
import openpyxl
from faker import Faker

#create a workbook from scratch
wb=Workbook()
wb.create_sheet("FakeExcel.xlsx")

#create a sheet inside workbook
wb.active.title="fake sheet"

#open the workbook
fakeExcel=openpyxl.load_workbook("FakeExcel.xlsx")
sh=fakeExcel["fake sheet"]

#create Fake object
fakeFile=Faker()

#insert fake data
for i in range (1,11):
        sh.cell(i,1).value=fakeFile.name()
        sh.cell(i,2).value=fakeFile.email()
        sh.cell(i,3).value=fakeFile.address()

fakeExcel.save("FakeExcel.xlsx")