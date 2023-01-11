from openpyxl import workbook
import openpyxl

wb=openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\test excels\\multiple rows.xlsx")
ws=wb.active
row_count=ws.max_row
col_count=ws.max_column
print("row count is" + " " + str(row_count))
# for i in range (1, row_count+1):
#     for cell in range (1, col_count+1):
#         print(ws.cell(i,cell).value)
for i in ws['A1':'C4']:
    for c in i:
        print(c.value)